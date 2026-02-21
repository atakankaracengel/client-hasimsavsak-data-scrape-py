import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import time
import re

class YokAtlasNetScraper:
    def __init__(self):
        self.base_url = "https://yokatlas.yok.gov.tr"
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def get_lisans_programs(self):
        """Lisans programlarını çeker"""
        print("Lisans programları çekiliyor...")
        url = f"{self.base_url}/netler.php"

        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')

            # Lisans programlarını bul
            select_element = soup.find('select', {'id': 'bolum'})

            if not select_element:
                print("Lisans programları bulunamadı!")
                return []

            programs = []
            for option in select_element.find_all('option'):
                value = option.get('value')
                text = option.text.strip()
                if value:
                    programs.append({
                        'kod': value,
                        'program_adi': text
                    })

            print(f"{len(programs)} adet lisans programı bulundu.")
            return programs

        except Exception as e:
            print(f"Hata oluştu: {e}")
            return []

    def extract_university_code(self, link_element):
        """Link elementinden üniversite kodunu çeker"""
        if link_element:
            href = link_element.get('href', '')
            # lisans.php?y=103390230 formatından 103390230'u çek
            match = re.search(r'y=(\d+)', href)
            if match:
                return match.group(1)
        return ''

    def scrape_program_data(self, program_code, program_name):
        """Belirli bir program için net verilerini çeker"""
        print(f"\n'{program_name}' programı işleniyor... (Kod: {program_code})")
        url = f"{self.base_url}/netler-tablo.php?b={program_code}"

        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')

            table = soup.find('table', {'id': 'mydata'})

            if not table:
                print(f"  Tablo bulunamadı!")
                return None

            # Başlıkları al
            headers = []
            thead = table.find('thead')
            if thead:
                header_row = thead.find('tr')
                th_list = header_row.find_all('th')

                for idx, th in enumerate(th_list):
                    header_text = th.text.strip()

                    # Boş başlıkları kontrol et - bu 7. sütundaki boş başlık
                    if (header_text == '' or header_text == ' ' or header_text == '\xa0' or header_text == '&nbsp;'):
                        # Eğer bir önceki başlık "Yerleşen Son Kişi" ise, bu katsayılı puan sütunudur
                        if idx > 0 and 'Yerleşen Son Kişi' in th_list[idx-1].text:
                            header_text = '0.12 Katsayı ile Yerleşen Son Kişinin Puanı'
                        else:
                            header_text = f'Boş_Sütun_{idx}'

                    # Eğer aynı isimli sütun varsa, sonuna sayı ekle
                    original_header = header_text
                    counter = 1
                    while header_text in headers:
                        header_text = f"{original_header}_{counter}"
                        counter += 1

                    headers.append(header_text)

            # Satırları al
            data_rows = []
            tbody = table.find('tbody')
            if tbody:
                for row in tbody.find_all('tr'):
                    cells = row.find_all('td')
                    row_data = []
                    university_code = ''

                    for idx, cell in enumerate(cells):
                        # Link içindeki metni al
                        link = cell.find('a')
                        if link:
                            text = link.text.strip()
                            # İlk link genelde üniversite linki, ondan kodu çek
                            if idx == 1:  # İkinci sütun genelde üniversite
                                university_code = self.extract_university_code(link)
                        else:
                            text = cell.text.strip()
                        row_data.append(text)

                    if len(row_data) > 1:  # Boş satırları atla
                        # Üniversite kodunu row_data'ya ekle (en başa)
                        row_data.insert(1, university_code)  # İlk boş sütundan sonra
                        data_rows.append(row_data)

            if not data_rows:
                print(f"  Veri bulunamadı!")
                return None

            # Üniversite Kodu başlığını ekle
            headers.insert(1, 'Üniversite Kodu')  # İlk boş sütundan sonra

            # Sütun sayısını kontrol et ve düzelt
            header_count = len(headers)
            for i, row in enumerate(data_rows):
                if len(row) != header_count:
                    # Eksik sütunları boş değerle doldur
                    while len(row) < header_count:
                        row.append('')
                    # Fazla sütunları kırp
                    if len(row) > header_count:
                        row = row[:header_count]
                    data_rows[i] = row

            # DataFrame oluştur
            df = pd.DataFrame(data_rows, columns=headers)

            # İlk sütun boşsa kaldır (genelde boş bir sütun oluyor)
            first_col_name = str(df.columns[0]).strip()
            if first_col_name.startswith('Boş_Sütun') or first_col_name in ['', ' ', '\xa0', '&nbsp;']:
                df = df.iloc[:, 1:]

            # Program bilgilerini ekle
            df.insert(0, 'Program Kodu', program_code)
            df.insert(0, 'Program Adı', program_name)

            print(f"  ✓ {len(df)} satır, {len(df.columns)} sütun")
            return df

        except Exception as e:
            print(f"  ✗ Hata: {e}")
            import traceback
            traceback.print_exc()
            return None

    def save_to_csv(self, all_data, filename='yokatlas_netler.csv'):
        """Verileri CSV'ye kaydet"""
        if not all_data:
            return None

        # Tüm sütunları birleştir
        combined_df = pd.concat(all_data, ignore_index=True, sort=False)

        # NaN değerleri boş string ile değiştir
        combined_df = combined_df.fillna('')

        combined_df.to_csv(filename, index=False, encoding='utf-8-sig')
        print(f"\n✓ CSV: {filename}")
        print(f"  - {len(combined_df)} satır")
        print(f"  - {len(combined_df.columns)} sütun")
        return combined_df

    def save_to_excel(self, df, filename='yokatlas_netler.xlsx'):
        """Verileri Excel'e kaydet"""
        if df is None or df.empty:
            return

        print(f"\n✓ Excel: {filename}")

        # NaN değerleri boş string ile değiştir
        df = df.fillna('')

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Netler', index=False)

            workbook = writer.book
            worksheet = writer.sheets['Netler']

            # Başlık formatı
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Sütun genişliklerini ayarla
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Hücreleri ortala
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    def run(self, limit=None):
        """Ana fonksiyon"""
        print("="*60)
        print("YÖK ATLAS NET VERİLERİ ÇEKME - LISANS")
        print("="*60)

        # Programları çek
        programs = self.get_lisans_programs()

        if not programs:
            print("❌ Program listesi alınamadı!")
            return

        # Limit varsa uygula
        if limit:
            programs = programs[:limit]
            print(f"\n⚠️  İlk {limit} program işlenecek (test modu)")

        all_data = []
        success_count = 0
        fail_count = 0

        # Her program için veri çek
        for i, program in enumerate(programs, 1):
            print(f"\n[{i}/{len(programs)}] ", end="")
            df = self.scrape_program_data(program['kod'], program['program_adi'])

            if df is not None:
                all_data.append(df)
                success_count += 1
            else:
                fail_count += 1

            # Bekle
            if i < len(programs):
                time.sleep(0.5)

        # Verileri kaydet
        if all_data:
            print("\n" + "="*60)
            print("KAYIT İŞLEMİ")
            print("="*60)

            combined_df = self.save_to_csv(all_data)

            if combined_df is not None:
                self.save_to_excel(combined_df)

            print("\n" + "="*60)
            print("✅ İŞLEM TAMAMLANDI!")
            print("="*60)
            print(f"Başarılı: {success_count} program")
            print(f"Başarısız: {fail_count} program")
        else:
            print("\n❌ Hiç veri çekilemedi!")

# ÇALIŞTIR
if __name__ == "__main__":
    scraper = YokAtlasNetScraper()

    # Test: İlk 5 program
    #scraper.run(limit=5)
    scraper.run()
    # Tümü için: scraper.run()
